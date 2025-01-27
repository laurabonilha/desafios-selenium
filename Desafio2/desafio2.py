'''
PROPOSTA DO DESAFIO:
CRIAR UMA AUTOMAÇÃO UTILIZANDO SELENIUM QUE CAPTURE DADOS DE UM ARQUIVO JSON COM NOMES DE USUÁRIOS QUE DEVEM SER PESQUISADOS EM UMA PÁGINA DA WEB
ESSA PÁGINA IRÁ RETORNAR DIVERSOS RESULTADOS DE USUÁRIOS (NOME, PROFISSÃO, EMAIL, TELEFONE, USUÁRIO, ESTADO E IMAGEM DE APRESENTAÇÃO)
A PÁGINA FOI CONSTRUÍDA PARA QUE APRESENTE DELAYS ALEATÓRIO NA APARIÇÃO DOS ELEMENTOS PARA PRATICAR FUNÇÕES DE ESPERA
ESSES DADOS DEVEM SER CAPTURADOS E ESCRITOS EM UMA PLANILHA DE EXCEL

'''

import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
import re

var_strArquivoExcel = "dados_pagina.xlsx"

driver = webdriver.Chrome()
driver.get('https://curso-web-scraping.pages.dev/#/desafio/2')
driver.maximize_window()

with open ('desafio_2.json', 'r', encoding='utf-8') as file:
    dados = json.load(file)

print (dados)

# Loop para capturar cada usuário a ser pesquisado

for usuario in dados:
    print(f"Processando usuário: {usuario}")
    var_inputUsuario = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/main/div[2]/div/div[2]/input')
    var_inputUsuario.clear()
    var_inputUsuario.send_keys(usuario)
    var_btnPesquisar = driver.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/main/div[2]/div/div[2]/button')
    var_btnPesquisar.click()

    # Criando uma variável de espera padrão
    wait = WebDriverWait(driver=driver, timeout=15, poll_frequency=1)

    # Espera aparecer o a tabela
    wait.until(EC.visibility_of_element_located(locator=(By.XPATH, '//*[@id="root"]/div/div[2]/main/div[2]/div/section/div/div[2]')))

    # Encontrando os usuários dentro da tabela
    var_listaUsuarios = driver.find_elements(By.CSS_SELECTOR, 'div.items-center.bg-gray-50')

    #Loop ára capturar os dados de todos da lista de usuários da página

    for usuario_atual in var_listaUsuarios:
        var_strTextoUsuario = usuario_atual.get_attribute('innerText')
        
        # Função auxiliar para aplicar regex com verificação
        def extract_with_regex(pattern, text, group_index=0):
            match = re.search(pattern, text)
            return match.group(group_index).strip() if match else None

        # Aplica Regex para capturar as informações
        lines = var_strTextoUsuario.split('\n')
    
        # Nome é sempre a primeira linha
        var_strNomeUsuario = lines[0].strip() if len(lines) > 0 else None
    
        # Cargo é sempre a segunda linha
        var_strCargoUsuario = lines[1].strip() if len(lines) > 1 else None
    
        # Aplica Regex para capturar os dados (sem os prefixos)
        var_strEmailUsuario = extract_with_regex(r'(?<=E-mail:\s)([^\n]+)', var_strTextoUsuario)  
        var_strTelefoneUsuario = extract_with_regex(r'(?<=Telefone:\s)([^\n]+)', var_strTextoUsuario)  
        var_strUsuarioUsuario = extract_with_regex(r'(?<=Usuário:\s)([^\n]+)', var_strTextoUsuario)  
        var_strEstadoUsuario = extract_with_regex(r'(?<=Estado:\s)([^\n]+)', var_strTextoUsuario) 

        # Escreve o item no arquivo Excel
        # Tenta encontrar a ativar a planilha
        try:
            workbook = load_workbook(var_strArquivoExcel)
            sheet = workbook.active
        # Se não existir, cria um novo arquivo
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            # Define a ordem e nome das colunas
            sheet.append(['NOME', 'CARGO', 'E-MAIL', 'TELEFONE', 'USUÁRIO', 'ESTADO'])
        
        # Escreve os dados do usuário processado atual e salva
        sheet.append([var_strNomeUsuario, var_strCargoUsuario, var_strEmailUsuario, var_strTelefoneUsuario, var_strUsuarioUsuario, var_strEstadoUsuario])
        workbook.save(var_strArquivoExcel)


        

        
